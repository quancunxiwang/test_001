import openpyxl
from config.config import root_path

class Read_File():
    def __init__(self):
        self.excel_path='data\data.xlsx'

    def read_excel(self):
        wb=openpyxl.load_workbook(self.excel_path)
        table=wb['Sheet1']
        col_max=table.max_column
        row_max=table.max_row

        res=[]
        for i in range(2,row_max+1):
            res1=[]
            for j in range(1,col_max+1):
                values=table.cell(row=i,column=j).value
                res1.append(values)
            res.append(res1)
        return res
    
if __name__=="__main__":
    demo=Read_File()
    print(demo.read_excel())